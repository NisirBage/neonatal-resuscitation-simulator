from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from uuid import UUID

from app.fsm import SimulationEvent
from app.scenario import Scenario
from app.services.metrics_service import compute_session_metrics

# All clinical wording lives here — this module is a pure consumer.
from app import clinical_script

# Re-export get_clinical_phase so existing call-sites don't break.
get_clinical_phase = clinical_script.get_clinical_phase


# ─────────────────────────────────────────────────────────────────────────────
# INTERNAL ROW ACCUMULATOR
# ─────────────────────────────────────────────────────────────────────────────

class _Row:
    __slots__ = (
        "voice_command",
        "student_response",
        "system_action",
        "instructor_action",
        "notes",
    )

    def __init__(self) -> None:
        self.voice_command = ""
        self.student_response = ""
        self.system_action = ""
        self.instructor_action = ""
        self.notes = ""

    def append_system(self, text: str) -> None:
        if self.system_action:
            self.system_action = f"{self.system_action}; {text}"
        else:
            self.system_action = text


@dataclass
class _XlsxRow:
    """Extended row for the XLSX report — includes clinical phase and row type."""
    time: str
    clinical_phase: str
    voice_prompt: str
    student_response: str
    system_action: str
    instructor_action: str
    notes: str
    is_timer_row: bool = field(default=False)
    is_instructor_row: bool = field(default=False)


# ─────────────────────────────────────────────────────────────────────────────
# SCENARIO HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _build_prompt_map(scenario: Scenario) -> dict[str, str]:
    """state_id → CSV prompt string, sourced exclusively from clinical_script."""
    return {state.id: clinical_script.get_csv_prompt(state.id) for state in scenario.states}


def _build_auto_timer_map(scenario: Scenario) -> dict[str, list[str]]:
    """state_id → list of 'X Timer Started' labels for auto-start timers."""
    result: dict[str, list[str]] = {}
    for state in scenario.states:
        labels = [
            clinical_script.get_timer_started_label(t.id)
            for t in state.timers
            if t.auto_start
        ]
        if labels:
            result[state.id] = labels
    return result


def _build_state_phase_map(scenario: Scenario) -> dict[str, str]:
    """state_id → clinical phase name."""
    return {state.id: clinical_script.get_clinical_phase(state.id) for state in scenario.states}


def _fmt_duration(secs: float) -> str:
    m, s = divmod(int(secs), 60)
    if m > 0:
        return f"{m} min {s} sec"
    return f"{s} sec"


# ─────────────────────────────────────────────────────────────────────────────
# CSV EXPORT
# ─────────────────────────────────────────────────────────────────────────────

_COLUMNS = [
    "Time",
    "Voice Command",
    "Student Response",
    "System Action",
    "Instructor Action",
    "Notes",
]


def _build_terminal_states(scenario: Scenario) -> set[str]:
    """State IDs that end the simulation — they have no subsequent voice command."""
    return {
        state.id
        for state in scenario.states
        if state.metadata.get("terminal") or not state.transitions
    }


def _build_decision_states(scenario: Scenario) -> set[str]:
    """State IDs that present a yes/no clinical decision.

    Determined from scenario action types — no hardcoding of state names.
    Any state with a yes_no action produces a YES/NO protocol row when the
    student responds.
    """
    return {
        state.id
        for state in scenario.states
        for action in state.actions
        if action.type == "yes_no"
    }


def generate_clinical_csv(
    session_id: UUID,
    history: list[SimulationEvent],
    scenario: Scenario,
) -> str:
    """
    Clock-driven clinical timeline CSV.

    Three explicit phases:

      Phase 1 — CLOCK
        Allocate one _Row slot for every elapsed second (None = no events yet).
        The clock is authoritative.

      Phase 2 — OVERLAY
        Walk the FSM event list once (O(N)) and annotate the appropriate second.

        Voice Command — driven by state ENTRY with YES/NO protocol rows:
          • session_started  → prompt for the initial state at second 0.
          • state_transition → prompt for the new (target) state at the
                               transition second.  If a YES/NO decision row
                               already occupies that second, the prompt is
                               deferred to second S+1.  Terminal states silent.
          • student_input / audio_input → "YES/NO" for any decision state
                               (state with a yes_no action type), so the
                               professor's required protocol row appears.

        Student Response — driven by student input only:
          • student_input / audio_input → writes "YES" or "NO" to the
                                          Student Response column at the
                                          answer second.

        System Action — outcomes and timer labels at the event second.
        Timer alarms  — at the timer_event second.
        Instructor    — at the instructor_event second.

      Phase 3 — EMIT
        Stream every second in order.  Unannotated seconds are blank.

    Phrase content sourced exclusively from clinical_script.
    Returns a UTF-8 BOM-prefixed CSV string.
    """
    if not history:
        return _empty_csv()

    prompt_map     = _build_prompt_map(scenario)
    auto_timer_map = _build_auto_timer_map(scenario)

    t0     = history[0].timestamp
    t_last = history[-1].timestamp
    duration_seconds = max(1, int((t_last - t0).total_seconds()) + 1)

    # ── Phase 1: CLOCK ────────────────────────────────────────────────────────
    timeline: list[_Row | None] = [None] * duration_seconds

    def _offset(ts: datetime) -> int:
        return min(max(0, int((ts - t0).total_seconds())), duration_seconds - 1)

    def _row_at_offset(s: int) -> _Row:
        s = min(max(0, s), duration_seconds - 1)
        if timeline[s] is None:
            timeline[s] = _Row()
        return timeline[s]  # type: ignore[return-value]

    terminal_states  = _build_terminal_states(scenario)
    decision_states  = _build_decision_states(scenario)

    # ── Phase 2: OVERLAY ──────────────────────────────────────────────────────
    for event in history:
        s = _offset(event.timestamp)

        if event.type == "session_started":
            row = _row_at_offset(s)
            row.voice_command = prompt_map.get(scenario.initial_state, "")
            row.append_system("Simulation Started")
            for label in auto_timer_map.get(scenario.initial_state, []):
                row.append_system(label)

        elif event.type == "state_transition":
            target = event.target_state_id or ""
            row = _row_at_offset(s)

            # System action (outcome, auto-timer labels) always lands on the
            # transition second.
            outcome = clinical_script.get_transition_outcome(target)
            if outcome:
                row.append_system(outcome)
            for label in auto_timer_map.get(target, []):
                row.append_system(label)

            # Voice prompt driven by state ENTRY.
            # If the current second already carries a YES/NO decision row (from
            # student_input), defer the new state's prompt to S+1 so both rows
            # appear.  Terminal states emit no voice command.
            if target not in terminal_states:
                prompt_s = s + 1 if row.voice_command else s
                if prompt_s < duration_seconds:
                    prompt_row = _row_at_offset(prompt_s)
                    if not prompt_row.voice_command:
                        prompt_row.voice_command = prompt_map.get(target, "")

        elif event.type == "student_input":
            row = _row_at_offset(s)
            if event.state_id in decision_states:
                row.voice_command = "YES/NO"
            row.student_response = str(event.payload.get("response", "")).upper()

        elif event.type == "audio_input":
            row = _row_at_offset(s)
            if event.state_id in decision_states:
                row.voice_command = "YES/NO"
            row.student_response = str(event.payload.get("transcript", "")).upper()

        elif event.type == "timer_event":
            timer_id = str(event.payload.get("timer_id", ""))
            _row_at_offset(s).append_system(clinical_script.get_timer_alarm_label(timer_id))

        elif event.type == "instructor_event":
            event_name = str(event.payload.get("event", ""))
            _row_at_offset(s).instructor_action = (
                clinical_script.get_instructor_action_label(event_name)
            )

        elif event.type == "no_transition":
            _row_at_offset(s).notes = "Input not recognised"

    # ── Phase 3: EMIT ─────────────────────────────────────────────────────────
    buffer = io.StringIO()
    buffer.write("﻿")  # UTF-8 BOM for Excel compatibility
    writer = csv.DictWriter(buffer, fieldnames=_COLUMNS, lineterminator="\n")
    writer.writeheader()

    for s in range(duration_seconds):
        wall = t0 + timedelta(seconds=s)
        row  = timeline[s]
        writer.writerow({
            "Time":              wall.strftime("%H:%M:%S"),
            "Voice Command":     row.voice_command     if row else "",
            "Student Response":  row.student_response  if row else "",
            "System Action":     row.system_action     if row else "",
            "Instructor Action": row.instructor_action if row else "",
            "Notes":             row.notes             if row else "",
        })

    return buffer.getvalue()


def _empty_csv() -> str:
    buffer = io.StringIO()
    buffer.write("﻿")
    writer = csv.DictWriter(buffer, fieldnames=_COLUMNS, lineterminator="\n")
    writer.writeheader()
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# XLSX EXPORT
# ─────────────────────────────────────────────────────────────────────────────

_XLSX_COLUMNS = [
    "Time",
    "Clinical Phase",
    "Voice Prompt",
    "Student Response",
    "System Action",
    "Instructor Action",
    "Notes",
]

_COL_VOICE   = 2  # C
_COL_STUDENT = 3  # D
_COL_SYSTEM  = 4  # E
_COL_INSTR   = 5  # F


def _build_xlsx_timeline(
    history: list[SimulationEvent],
    scenario: Scenario,
) -> list[_XlsxRow]:
    """O(N + D) — one pass over events, one pass over seconds."""
    if not history:
        return []

    prompt_map     = _build_prompt_map(scenario)
    auto_timer_map = _build_auto_timer_map(scenario)
    phase_map      = _build_state_phase_map(scenario)

    t0 = history[0].timestamp
    t_last = history[-1].timestamp
    duration_seconds = max(1, int((t_last - t0).total_seconds()) + 1)

    class _Acc:
        __slots__ = (
            "voice_prompt", "student_response", "system_action",
            "instructor_action", "notes", "clinical_phase",
            "is_timer_row", "is_instructor_row",
        )

        def __init__(self) -> None:
            self.voice_prompt = ""
            self.student_response = ""
            self.system_action = ""
            self.instructor_action = ""
            self.notes = ""
            self.clinical_phase = ""
            self.is_timer_row = False
            self.is_instructor_row = False

        def append_system(self, text: str) -> None:
            if self.system_action:
                self.system_action = f"{self.system_action}; {text}"
            else:
                self.system_action = text

    acc: dict[int, _Acc] = {}

    def _get(ts: datetime) -> _Acc:
        offset = max(0, int((ts - t0).total_seconds()))
        if offset not in acc:
            acc[offset] = _Acc()
        return acc[offset]

    for event in history:
        row = _get(event.timestamp)

        if event.type == "session_started":
            initial = scenario.initial_state
            row.voice_prompt   = prompt_map.get(initial, "")
            row.clinical_phase = phase_map.get(initial, clinical_script.get_clinical_phase(initial))
            row.append_system("Simulation Started")
            for label in auto_timer_map.get(initial, []):
                row.append_system(label)
                row.is_timer_row = True

        elif event.type == "state_transition":
            target = event.target_state_id or ""
            row.voice_prompt   = prompt_map.get(target, "")
            row.clinical_phase = phase_map.get(target, clinical_script.get_clinical_phase(target))
            for label in auto_timer_map.get(target, []):
                row.append_system(label)
                row.is_timer_row = True
            outcome = clinical_script.get_transition_outcome(target)
            if outcome:
                row.append_system(outcome)

        elif event.type == "student_input":
            resp = event.payload.get("response", "")
            row.student_response = str(resp).upper()
            if not row.clinical_phase:
                row.clinical_phase = phase_map.get(
                    event.state_id, clinical_script.get_clinical_phase(event.state_id)
                )

        elif event.type == "audio_input":
            transcript = event.payload.get("transcript", "")
            row.student_response = str(transcript).upper()
            if not row.clinical_phase:
                row.clinical_phase = phase_map.get(
                    event.state_id, clinical_script.get_clinical_phase(event.state_id)
                )

        elif event.type == "timer_event":
            timer_id = str(event.payload.get("timer_id", ""))
            row.append_system(clinical_script.get_timer_alarm_label(timer_id, xlsx=True))
            row.is_timer_row = True
            if not row.clinical_phase:
                row.clinical_phase = phase_map.get(
                    event.state_id, clinical_script.get_clinical_phase(event.state_id)
                )

        elif event.type == "instructor_event":
            event_name = str(event.payload.get("event", ""))
            row.instructor_action = clinical_script.get_instructor_action_label(event_name)
            row.is_instructor_row = True
            if not row.clinical_phase:
                row.clinical_phase = phase_map.get(
                    event.state_id, clinical_script.get_clinical_phase(event.state_id)
                )

        elif event.type == "no_transition":
            row.notes = "Student response not recognised"
            if not row.clinical_phase:
                row.clinical_phase = phase_map.get(
                    event.state_id, clinical_script.get_clinical_phase(event.state_id)
                )

    result: list[_XlsxRow] = []
    last_phase = clinical_script.get_clinical_phase(scenario.initial_state)

    for s in range(duration_seconds):
        wall = t0 + timedelta(seconds=s)
        a = acc.get(s)
        if a:
            phase = a.clinical_phase or last_phase
            last_phase = phase
            result.append(_XlsxRow(
                time=wall.strftime("%H:%M:%S"),
                clinical_phase=phase,
                voice_prompt=a.voice_prompt,
                student_response=a.student_response,
                system_action=a.system_action,
                instructor_action=a.instructor_action,
                notes=a.notes,
                is_timer_row=a.is_timer_row,
                is_instructor_row=a.is_instructor_row,
            ))
        else:
            result.append(_XlsxRow(
                time=wall.strftime("%H:%M:%S"),
                clinical_phase=last_phase,
                voice_prompt="",
                student_response="",
                system_action="",
                instructor_action="",
                notes="",
            ))

    return result


def generate_clinical_xlsx(
    session_id: UUID,
    history: list[SimulationEvent],
    scenario: Scenario,
    current_state_id: str,
) -> bytes:
    """
    Generate a professional instructor assessment report as an XLSX workbook.
    Complexity: O(N + D). Returns raw bytes.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    c_dark_blue  = "1F4E79"
    c_grey_bg    = "EEEEEE"
    c_voice      = "D6E4F0"
    c_student    = "D5F5E3"
    c_system     = "FAE5D3"
    c_instructor = "FADBD8"
    c_timer      = "FFFDE7"
    c_white      = "FFFFFF"

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
        return Font(bold=bold, color=color, size=size, name="Calibri")

    wrap_align   = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    from app.services.report_service import compute_training_score

    metrics = compute_session_metrics(history=history, current_state_id=current_state_id)
    score   = compute_training_score(
        no_transition_count=metrics["no_transition_count"],
        instructor_intervention_count=metrics["instructor_intervention_count"],
    )

    timeline = _build_xlsx_timeline(history, scenario)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clinical Timeline"

    n_xlsx_cols = len(_XLSX_COLUMNS)

    def _merge_header(row: int, text: str, fill_hex: str,
                      font_color: str = "FFFFFF", size: int = 12, bold: bool = True) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_xlsx_cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill  = _fill(fill_hex)
        cell.font  = _font(bold=bold, color=font_color, size=size)
        cell.alignment = center_align
        ws.row_dimensions[row].height = 22

    def _meta_row(row: int, label: str, value: str) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = _fill(c_grey_bg)
        lc.font = _font(bold=True, color="333333")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=row, column=2, value=value)
        vc.fill = _fill(c_grey_bg)
        vc.font = _font(bold=False, color="000000")
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_xlsx_cols)
        ws.row_dimensions[row].height = 16

    current_row = 1
    _merge_header(current_row, "NEONATAL RESUSCITATION SIMULATION REPORT",
                  c_dark_blue, "FFFFFF", size=14, bold=True)
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=n_xlsx_cols)
    ws.row_dimensions[current_row].height = 6
    current_row += 1

    t0     = history[0].timestamp if history else datetime.now(timezone.utc)
    t_last = history[-1].timestamp if history else t0
    duration_secs = (t_last - t0).total_seconds()
    sim_date   = t0.strftime("%d %B %Y")
    start_time = t0.strftime("%H:%M:%S UTC")
    end_time   = t_last.strftime("%H:%M:%S UTC") if history else "—"
    outcome    = clinical_script.get_outcome_label(current_state_id)

    _meta_row(current_row, "Session ID",      str(session_id)); current_row += 1
    _meta_row(current_row, "Scenario Name",   scenario.name);   current_row += 1
    _meta_row(current_row, "Simulation Date", sim_date);        current_row += 1
    _meta_row(current_row, "Start Time",      start_time);      current_row += 1
    _meta_row(current_row, "Completion Time", end_time);        current_row += 1
    _meta_row(current_row, "Total Duration",  _fmt_duration(duration_secs)); current_row += 1
    _meta_row(current_row, "Outcome",         outcome);         current_row += 1
    _meta_row(current_row, "Training Score",  f"{score} / 100"); current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=n_xlsx_cols)
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    header_row = current_row
    for col_idx, col_name in enumerate(_XLSX_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = _fill(c_dark_blue)
        cell.font = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = center_align
    ws.row_dimensions[header_row].height = 20
    current_row += 1

    ws.freeze_panes = ws.cell(row=current_row, column=1)

    for trow in timeline:
        is_timer = trow.is_timer_row
        values = [
            trow.time,
            trow.clinical_phase,
            trow.voice_prompt,
            trow.student_response,
            trow.system_action,
            trow.instructor_action,
            trow.notes,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.alignment = wrap_align
            cell.font = _font()
            if is_timer:
                cell.fill = _fill(c_timer)
            else:
                if col_idx - 1 == _COL_VOICE   and val: cell.fill = _fill(c_voice)
                elif col_idx - 1 == _COL_STUDENT and val: cell.fill = _fill(c_student)
                elif col_idx - 1 == _COL_SYSTEM  and val: cell.fill = _fill(c_system)
                elif col_idx - 1 == _COL_INSTR   and val: cell.fill = _fill(c_instructor)
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=n_xlsx_cols)
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    _merge_header(current_row, "SIMULATION SUMMARY", c_grey_bg, "333333", size=11, bold=True)
    current_row += 1

    def _summary_row(label: str, value: str) -> None:
        nonlocal current_row
        lc = ws.cell(row=current_row, column=1, value=label)
        lc.fill = _fill(c_grey_bg)
        lc.font = _font(bold=True, color="333333")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=current_row, column=2, value=value)
        vc.fill = _fill(c_grey_bg)
        vc.font = _font(bold=False)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=current_row, start_column=2,
                       end_row=current_row, end_column=n_xlsx_cols)
        ws.row_dimensions[current_row].height = 16
        current_row += 1

    _summary_row("Outcome",                  outcome)
    _summary_row("Duration",                 _fmt_duration(duration_secs))
    _summary_row("Voice Responses",          str(metrics["student_input_count"]))
    _summary_row("Incorrect Responses",      str(metrics["no_transition_count"]))
    _summary_row("Instructor Interventions", str(metrics["instructor_intervention_count"]))
    _summary_row("Timer Events",             str(metrics["timer_event_count"]))
    _summary_row("Successful Transitions",   str(metrics["successful_transition_count"]))
    _summary_row("Training Score",           f"{score} / 100")

    col_widths = [len(h) + 4 for h in _XLSX_COLUMNS]
    for trow in timeline:
        for i, v in enumerate([
            trow.time, trow.clinical_phase, trow.voice_prompt,
            trow.student_response, trow.system_action,
            trow.instructor_action, trow.notes,
        ]):
            col_widths[i] = max(col_widths[i], min(len(str(v)), 60))

    min_widths = [10, 22, 18, 16, 20, 22, 18]
    for col_idx, (width, min_w) in enumerate(zip(col_widths, min_widths), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, min_w)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
