"""
20 test cases for clinical_timeline_service.

Tests 1–10: generate_clinical_csv (original, backward-compatible)
Tests 11–20: generate_clinical_xlsx + enhanced service features

Run from the backend/ directory:
    pytest tests/test_clinical_timeline_service.py -v
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta, timezone
from uuid import uuid4

import pytest

from app.fsm import SimulationEvent
from app.scenario import Action, Scenario, State, Timer, Transition
from app.services.clinical_timeline_service import (
    generate_clinical_csv,
    generate_clinical_xlsx,
    get_clinical_phase,
)

# ── Fixtures / helpers ─────────────────────────────────────────────────────────

_BASE = datetime(2024, 1, 1, 10, 30, 0, tzinfo=timezone.utc)


def _ts(offset: float) -> datetime:
    return _BASE + timedelta(seconds=offset)


def _event(
    type: str,
    state_id: str,
    offset: float = 0.0,
    payload: dict | None = None,
    target_state_id: str | None = None,
) -> SimulationEvent:
    return SimulationEvent(
        type=type,
        state_id=state_id,
        timestamp=_ts(offset),
        payload=payload or {},
        target_state_id=target_state_id,
    )


def _make_scenario() -> Scenario:
    """Minimal valid two-state scenario with a yes_no action and auto-start timer."""
    return Scenario(
        id="test_scenario",
        name="Test Scenario",
        version="1.0.0",
        initial_state="baby_born",
        states=[
            State(
                id="baby_born",
                name="Baby Born",
                metadata={"voice_prompt": "Baby is born. Confirm when ready."},
                actions=[
                    Action(id="confirm_birth", type="yes_no", options=["yes", "no"]),
                ],
                timers=[
                    Timer(
                        id="birth_timer",
                        duration_seconds=60,
                        event="birth_timer_elapsed",
                        auto_start=True,
                    ),
                ],
                transitions=[
                    Transition(
                        id="t_yes",
                        trigger="action",
                        target_state="crying_assessment",
                        action_id="confirm_birth",
                        expected_response="yes",
                    ),
                    Transition(
                        id="t_no",
                        trigger="action",
                        target_state="crying_assessment",
                        action_id="confirm_birth",
                        expected_response="no",
                    ),
                    Transition(
                        id="t_timer",
                        trigger="timer",
                        target_state="crying_assessment",
                        timer_id="birth_timer",
                    ),
                ],
            ),
            State(
                id="crying_assessment",
                name="Is Baby Crying?",
                metadata={"voice_prompt": "Is the baby crying?"},
                actions=[
                    Action(id="is_crying", type="yes_no", options=["yes", "no"]),
                ],
                timers=[],
                transitions=[
                    Transition(
                        id="t_cry_yes",
                        trigger="action",
                        target_state="routine_care",
                        action_id="is_crying",
                        expected_response="yes",
                    ),
                    Transition(
                        id="t_cry_no",
                        trigger="action",
                        target_state="simulation_complete",
                        action_id="is_crying",
                        expected_response="no",
                    ),
                ],
            ),
            State(
                id="routine_care",
                name="Routine Care",
                metadata={"voice_prompt": "Perform routine care.", "terminal": True},
                actions=[],
                timers=[],
                transitions=[],
            ),
            State(
                id="simulation_complete",
                name="Simulation Complete",
                metadata={"voice_prompt": "Resuscitation complete.", "terminal": True},
                actions=[],
                timers=[],
                transitions=[],
            ),
        ],
    )


def _parse_rows(csv_content: str) -> list[dict]:
    """Strip BOM and parse CSV into list of dicts."""
    content = csv_content.lstrip("﻿")
    reader = csv.DictReader(io.StringIO(content))
    return list(reader)


# ── Test 1: Empty history returns header only ──────────────────────────────────

def test_empty_history_returns_header_only() -> None:
    result = generate_clinical_csv(uuid4(), [], _make_scenario())

    assert result.startswith("﻿"), "Missing UTF-8 BOM"
    rows = _parse_rows(result)
    assert rows == [], f"Expected no data rows, got {len(rows)}"


# ── Test 2: UTF-8 BOM and correct column headers ───────────────────────────────

def test_bom_and_correct_columns() -> None:
    history = [_event("session_started", "baby_born", 0)]
    result = generate_clinical_csv(uuid4(), history, _make_scenario())

    assert result.startswith("﻿")
    rows = _parse_rows(result)
    for col in ["Time", "Voice Command / Response", "System Action", "Instructor Action", "Notes"]:
        assert col in rows[0], f"Column '{col}' missing"
    assert "Student Response" not in rows[0], "Student Response column must not appear in CSV"


# ── Test 3: session_started sets voice prompt + system action + auto timer ─────

def test_session_started_row() -> None:
    history = [_event("session_started", "baby_born", 0)]
    result = generate_clinical_csv(uuid4(), history, _make_scenario())

    rows = _parse_rows(result)
    # The Time column carries the clock; Voice Command is phrase-only, no timestamp.
    assert rows[0]["Voice Command / Response"] == "BABY BORN AT"
    assert "Simulation Started" in rows[0]["System Action"]
    assert "BIRTH TIMER STARTED" in rows[0]["System Action"]


# ── Test 4: Decision states show YES/NO in Voice Command / Response column ─────

def test_decision_states_show_yes_no_prompt() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("student_input", "baby_born", 3, {"action_id": "confirm_birth", "response": "yes"}),
        _event("state_transition", "baby_born", 3, target_state_id="crying_assessment"),
        _event("student_input", "crying_assessment", 7, {"action_id": "is_crying", "response": "no"}),
        _event("state_transition", "crying_assessment", 7, target_state_id="simulation_complete"),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    # Student responses appear as the actual answer in the unified column.
    assert rows[3]["Voice Command / Response"] == "YES"
    assert rows[7]["Voice Command / Response"] == "NO"


# ── Test 5: YES/NO protocol rows + deferred state prompt when simultaneous ──────

def test_yes_no_protocol_rows_restored() -> None:
    """
    When student answers and FSM transitions on the same second:
      - YES/NO appears in Voice Command (decision protocol row)
      - Student Response carries the actual answer
      - New state's prompt deferred to S+1 (reactive, no pre-scan)
    """
    history = [
        _event("session_started", "baby_born", 0),
        _event("student_input", "baby_born", 5, {"response": "yes"}),
        _event("state_transition", "baby_born", 5, target_state_id="crying_assessment"),
        _event("student_input", "crying_assessment", 9, {"response": "yes"}),
        _event("state_transition", "crying_assessment", 9, target_state_id="routine_care"),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    # Second 5: student answered — actual response in Voice Command / Response.
    assert rows[5]["Voice Command / Response"] == "YES"
    # New state's prompt deferred to S+1 because the response occupies second 5.
    assert rows[6]["Voice Command / Response"] == "IS BABY CRYING"

    # Second 9: terminal transition — actual response; ROUTINE CARE system action.
    assert rows[9]["Voice Command / Response"] == "YES"
    assert rows[9]["System Action"]            == "ROUTINE CARE"


# ── Test 6: Timer alarm shows clinical label ───────────────────────────────────

def test_timer_alarm_shows_clinical_label() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("timer_event", "baby_born", 60, {"timer_id": "birth_timer"}),
        _event("state_transition", "baby_born", 60, target_state_id="crying_assessment"),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    assert len(rows) == 61  # 0..60 inclusive
    assert "BIRTH TIMER ALARM" in rows[60]["System Action"]


# ── Test 7: Multiple timer alarms in same second both appear ──────────────────

def test_multiple_timer_alarms_concatenated() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("timer_event", "baby_born", 5, {"timer_id": "birth_timer"}),
        _event("timer_event", "baby_born", 5, {"timer_id": "ventilation_timer"}),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    system = rows[5]["System Action"]
    assert "BIRTH TIMER ALARM" in system
    assert "VENTILATION TIMER ALARM" in system


# ── Test 8: Instructor event shows formatted action name ──────────────────────

def test_instructor_event_formatted() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("instructor_event", "baby_born", 4, {"event": "baby_not_crying"}),
        _event("state_transition", "baby_born", 4, target_state_id="crying_assessment"),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    assert rows[4]["Instructor Action"] == "Instructor forced transition"


# ── Test 9: no_transition adds note, does not crash ───────────────────────────

def test_no_transition_adds_note() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("student_input", "baby_born", 2, {"response": "maybe"}),
        _event("no_transition", "baby_born", 2),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    assert "not recognised" in rows[2]["Notes"].lower()


# ── Test 10: Time column format and total row count ───────────────────────────

def test_time_column_and_row_count() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("state_transition", "baby_born", 65, target_state_id="routine_care"),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    assert len(rows) == 66  # 0..65 inclusive
    assert rows[0]["Time"] == _BASE.astimezone().strftime("%H:%M:%S")
    assert rows[65]["Time"] == (_BASE + timedelta(seconds=65)).astimezone().strftime("%H:%M:%S")


# ── Test 10b: 20-second wait — blank rows preserved, columns independent ──────

def test_twenty_second_wait_blank_rows() -> None:
    """
    Trainee waits 20 seconds before answering.  The CSV must contain:
      - voice prompt at the second the state was ENTERED (second 7)
      - blank rows while the trainee waits (seconds 8-26)
      - student response + system action on the answer second (second 27)
      - voice command at second 27 is blank (routine_care is terminal)
    """
    history = [
        _event("session_started", "baby_born", 0),
        _event("state_transition", "baby_born", 7, target_state_id="crying_assessment"),
        _event("student_input", "crying_assessment", 27, {"response": "yes"}),
        _event("state_transition", "crying_assessment", 27, target_state_id="routine_care"),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    # State entered at second 7 — voice command appears here, not at answer time.
    assert rows[7]["Voice Command / Response"] == "IS BABY CRYING"

    # Seconds 8–26: completely blank
    for s in range(8, 27):
        assert rows[s]["Voice Command / Response"] == "", f"Row {s} voice should be blank"
        assert rows[s]["System Action"]            == "", f"Row {s} system should be blank"

    # Second 27: actual student response + system action.
    assert rows[27]["Voice Command / Response"] == "YES"
    assert rows[27]["System Action"]            == "ROUTINE CARE"


# ── Test 10c: terminal-state transitions never emit a voice command row ───────

def test_terminal_transition_no_voice_command() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("student_input", "baby_born", 3, {"response": "yes"}),
        _event("state_transition", "baby_born", 3, target_state_id="routine_care"),
    ]
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, _make_scenario()))

    # Second 3: actual student response; routine_care is terminal so no state prompt.
    assert rows[3]["Voice Command / Response"] == "YES"
    assert rows[3]["System Action"]            == "ROUTINE CARE"


# ── Test 10d: corrective-ventilation loop emits repeated prompts ──────────────

def _make_ventilation_scenario() -> Scenario:
    """Three-state loop: ventilation_in_progress → hr_check → corrective → hr_check …"""
    return Scenario(
        id="vent_loop",
        name="Ventilation Loop",
        version="1.0.0",
        initial_state="ventilation_in_progress",
        states=[
            State(
                id="ventilation_in_progress",
                name="Ventilation In Progress",
                metadata={"voice_prompt": "Measure heart rate every 15 seconds."},
                actions=[Action(id="ack", type="yes_no", options=["yes", "no"])],
                timers=[Timer(id="ventilation_timer", duration_seconds=30,
                              event="vent_elapsed", auto_start=True)],
                transitions=[
                    Transition(id="t1", trigger="action", target_state="heart_rate_after_ventilation",
                               action_id="ack", expected_response="yes"),
                    Transition(id="t2", trigger="timer", target_state="heart_rate_after_ventilation",
                               timer_id="ventilation_timer"),
                ],
            ),
            State(
                id="heart_rate_after_ventilation",
                name="Is HR Increasing?",
                metadata={"voice_prompt": "Is the heart rate increasing?"},
                actions=[Action(id="hr", type="yes_no", options=["yes", "no"])],
                timers=[],
                transitions=[
                    Transition(id="t3", trigger="action", target_state="ventilation_corrective_steps",
                               action_id="hr", expected_response="no"),
                    Transition(id="t4", trigger="action", target_state="simulation_complete",
                               action_id="hr", expected_response="yes"),
                ],
            ),
            State(
                id="ventilation_corrective_steps",
                name="Corrective Steps",
                metadata={"voice_prompt": "Follow ventilation corrective steps."},
                actions=[Action(id="done", type="yes_no", options=["yes", "no"])],
                timers=[Timer(id="corrective_ventilation_timer", duration_seconds=30,
                              event="corr_elapsed", auto_start=True)],
                transitions=[
                    Transition(id="t5", trigger="action", target_state="heart_rate_after_ventilation",
                               action_id="done", expected_response="yes"),
                    Transition(id="t6", trigger="timer", target_state="heart_rate_after_ventilation",
                               timer_id="corrective_ventilation_timer"),
                ],
            ),
            State(
                id="simulation_complete",
                name="Simulation Complete",
                metadata={"terminal": True, "voice_prompt": "Stop resuscitation."},
                actions=[], timers=[], transitions=[],
            ),
        ],
    )


def test_corrective_ventilation_loop_repeated_prompts() -> None:
    """
    Two corrective-loop iterations must produce complete YES/NO + prompt sequences.

    When student answers and FSM transitions simultaneously:
      - YES/NO appears at the answer second
      - New state prompt deferred to S+1 (reactive deferral, no pre-scan)
    """
    history = [
        _event("session_started", "ventilation_in_progress", 0),
        # Loop 1: timer fires at 30s → heart_rate_after_ventilation
        _event("timer_event", "ventilation_in_progress", 30, {"timer_id": "ventilation_timer"}),
        _event("state_transition", "ventilation_in_progress", 30,
               target_state_id="heart_rate_after_ventilation"),
        # Student answers NO at 35 → corrective steps
        _event("student_input", "heart_rate_after_ventilation", 35, {"response": "no"}),
        _event("state_transition", "heart_rate_after_ventilation", 35,
               target_state_id="ventilation_corrective_steps"),
        # Corrective timer fires at 65s → back to heart_rate_after_ventilation
        _event("timer_event", "ventilation_corrective_steps", 65,
               {"timer_id": "corrective_ventilation_timer"}),
        _event("state_transition", "ventilation_corrective_steps", 65,
               target_state_id="heart_rate_after_ventilation"),
        # Loop 2: student answers NO again at 70 → corrective steps
        _event("student_input", "heart_rate_after_ventilation", 70, {"response": "no"}),
        _event("state_transition", "heart_rate_after_ventilation", 70,
               target_state_id="ventilation_corrective_steps"),
        # Final: student answers YES at 75 → simulation_complete
        _event("student_input", "ventilation_corrective_steps", 75, {"response": "yes"}),
        _event("state_transition", "ventilation_corrective_steps", 75,
               target_state_id="simulation_complete"),
    ]
    scenario = _make_ventilation_scenario()
    rows = _parse_rows(generate_clinical_csv(uuid4(), history, scenario))

    # Second 0: initial state entered
    assert rows[0]["Voice Command / Response"] == "MEASURE HEART RATE EVERY 15 SECONDS"

    # Second 30: timer fires AND state entered — both on same row
    assert "VENTILATION TIMER ALARM" in rows[30]["System Action"]
    assert rows[30]["Voice Command / Response"] == "IS THE HEART RATE INCREASING"

    # Second 35: actual response NO; corrective steps prompt deferred to 36.
    assert rows[35]["Voice Command / Response"] == "NO"
    assert "CORRECTIVE VENTILATION TIMER 30 SEC" in rows[35]["System Action"]
    assert rows[36]["Voice Command / Response"] == "FOLLOW VENTILATION CORRECTIVE STEPS"

    # Second 65: corrective timer fires AND hr state re-entered
    assert "CORRECTIVE VENTILATION TIMER ALARM" in rows[65]["System Action"]
    assert rows[65]["Voice Command / Response"] == "IS THE HEART RATE INCREASING"

    # Second 70: loop 2 — actual response NO; corrective steps prompt deferred to 71.
    assert rows[70]["Voice Command / Response"] == "NO"
    assert "CORRECTIVE VENTILATION TIMER 30 SEC" in rows[70]["System Action"]
    assert rows[71]["Voice Command / Response"] == "FOLLOW VENTILATION CORRECTIVE STEPS"

    # Second 75: actual response YES; simulation_complete is terminal.
    assert rows[75]["Voice Command / Response"] == "YES"
    assert rows[75]["System Action"]            == "STOP RESUSCITATION PROTOCOL"


# ═════════════════════════════════════════════════════════════════════════════
# Tests 11–20: generate_clinical_xlsx + enhanced service helpers
# ═════════════════════════════════════════════════════════════════════════════

def _parse_xlsx(xlsx_bytes: bytes):
    """Return the active worksheet from XLSX bytes."""
    import io as _io
    import openpyxl
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes))
    return wb.active


def _xlsx_cell_values(ws) -> list[list]:
    """Extract all non-None rows from a worksheet as a list of lists."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            rows.append(list(row))
    return rows


# ── Test 11: Clinical phase mapping helper ────────────────────────────────────

def test_clinical_phase_known_states() -> None:
    assert get_clinical_phase("baby_born")                    == "Birth"
    assert get_clinical_phase("crying_assessment")            == "Respiratory Assessment"
    assert get_clinical_phase("heart_rate_assessment")        == "Heart Rate Assessment"
    assert get_clinical_phase("ventilation_in_progress")      == "Ventilation"
    assert get_clinical_phase("ventilation_corrective_steps") == "Corrective Ventilation"
    assert get_clinical_phase("routine_care")                 == "Routine Care"
    assert get_clinical_phase("simulation_complete")          == "Simulation Complete"


# ── Test 12: Unknown state falls back to title-cased state ID ─────────────────

def test_clinical_phase_unknown_state_fallback() -> None:
    result = get_clinical_phase("some_custom_state")
    assert result == "Some Custom State"


# ── Test 13: XLSX output is valid openpyxl workbook (non-empty bytes) ─────────

def test_xlsx_returns_valid_workbook() -> None:
    history = [_event("session_started", "baby_born", 0)]
    xlsx_bytes = generate_clinical_xlsx(
        uuid4(), history, _make_scenario(), "baby_born"
    )
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0
    ws = _parse_xlsx(xlsx_bytes)
    assert ws is not None
    assert ws.title == "Clinical Timeline"


# ── Test 14: Report header contains scenario name and session ID ───────────────

def test_xlsx_header_contains_metadata() -> None:
    sid = uuid4()
    history = [_event("session_started", "baby_born", 0)]
    xlsx_bytes = generate_clinical_xlsx(sid, history, _make_scenario(), "baby_born")
    ws = _parse_xlsx(xlsx_bytes)

    all_text = " ".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert str(sid) in all_text
    assert "Test Scenario" in all_text


# ── Test 15: XLSX contains "Clinical Phase" column header ─────────────────────

def test_xlsx_has_clinical_phase_column() -> None:
    history = [_event("session_started", "baby_born", 0)]
    xlsx_bytes = generate_clinical_xlsx(
        uuid4(), history, _make_scenario(), "baby_born"
    )
    ws = _parse_xlsx(xlsx_bytes)

    all_values = {
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "Clinical Phase" in all_values


# ── Test 16: XLSX summary section contains training score ─────────────────────

def test_xlsx_summary_contains_score() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("student_input", "baby_born", 3, {"response": "yes"}),
        _event("state_transition", "baby_born", 3, target_state_id="routine_care"),
    ]
    xlsx_bytes = generate_clinical_xlsx(
        uuid4(), history, _make_scenario(), "routine_care"
    )
    ws = _parse_xlsx(xlsx_bytes)

    all_text = " ".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Training Score" in all_text
    assert "100" in all_text  # perfect score — no errors


# ── Test 17: XLSX instructor action uses clinical language (not raw event name) ─

def test_xlsx_instructor_action_clinical_language() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("instructor_event", "baby_born", 4, {"event": "baby_not_crying"}),
        _event("state_transition", "baby_born", 4, target_state_id="crying_assessment"),
    ]
    xlsx_bytes = generate_clinical_xlsx(
        uuid4(), history, _make_scenario(), "crying_assessment"
    )
    ws = _parse_xlsx(xlsx_bytes)

    all_text = " ".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    # Clinical phrasing — NOT raw event name like "baby_not_crying"
    assert "Instructor" in all_text
    assert "baby_not_crying" not in all_text


# ── Test 18: Timer rows carry clinical reminder wording ───────────────────────

def test_xlsx_timer_reminder_wording() -> None:
    history = [
        _event("session_started", "baby_born", 0),
        _event("timer_event", "baby_born", 60, {"timer_id": "birth_timer"}),
        _event("state_transition", "baby_born", 60, target_state_id="crying_assessment"),
    ]
    xlsx_bytes = generate_clinical_xlsx(
        uuid4(), history, _make_scenario(), "crying_assessment"
    )
    ws = _parse_xlsx(xlsx_bytes)

    all_text = " ".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    # XLSX uses "Reminder" language, not "Alarm"
    assert "Reminder" in all_text
    assert "Birth Timer Alarm" not in all_text  # raw label must NOT appear


# ── Test 19: CSV remains byte-for-byte identical after service enhancement ─────

def test_csv_output_unchanged_after_xlsx_addition() -> None:
    """Confirm generate_clinical_csv() contract has not changed."""
    history = [
        _event("session_started", "baby_born", 0),
        _event("student_input", "baby_born", 3, {"response": "yes"}),
        _event("state_transition", "baby_born", 3, target_state_id="crying_assessment"),
    ]
    result = generate_clinical_csv(uuid4(), history, _make_scenario())

    # BOM still present
    assert result.startswith("﻿")
    # Clinical Phase column must NOT appear in CSV
    assert "Clinical Phase" not in result
    # Student Response column removed — unified into Voice Command / Response
    assert "Student Response" not in result
    # Correct columns present
    rows = _parse_rows(result)
    for col in ["Time", "Voice Command / Response", "System Action",
                "Instructor Action", "Notes"]:
        assert col in rows[0]


# ── Test 20: XLSX column headers frozen and columns autosized > 0 ─────────────

def test_xlsx_freeze_and_column_widths() -> None:
    history = [_event("session_started", "baby_born", 0)]
    xlsx_bytes = generate_clinical_xlsx(
        uuid4(), history, _make_scenario(), "baby_born"
    )
    ws = _parse_xlsx(xlsx_bytes)

    # Freeze panes should be set (not None)
    assert ws.freeze_panes is not None

    # All 7 columns should have explicit widths > 0
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(["Time","Clinical Phase","Voice Prompt",
                                   "Student Response","System Action",
                                   "Instructor Action","Notes"]) + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        assert dim is not None, f"Column {letter} has no dimension set"
        assert dim.width > 0, f"Column {letter} width is zero"
